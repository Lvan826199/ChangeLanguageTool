from openpyxl import load_workbook
from ChangeLanguageScript.changeLanguage_Test4.readExcelFile import RWExcel
#写入已存在的xlsx文件第一种方法
class Write_excel(object):
    '''修改excel数据'''
    def __init__(self,filename,index,msg):
        """

        :param filename: 测试的文件名
        :param index: 对应要写入的行数的下标
        :param msg: 需要写入的内容
        """
        self.filename = filename
        self.index = index
        self.msg = msg


    def write(self):
        #写入已存在的xlsx文件
        wb = load_workbook(self.filename)#生成一个已存在的wookbook对象
        wb1 = wb.active#激活sheet
        wb1.cell(self.index+1,5,self.msg)#往sheet中的第二行第五列写入msg的数据，注意这里的下标是从1开始
        wb.save(self.filename)  # 保存


    def write_result(self):
        result = RWExcel().differ(self.index)
        # 写入已存在的xlsx文件
        wb = load_workbook(self.filename)  # 生成一个已存在的wookbook对象
        wb1 = wb.active  # 激活sheet
        wb1.cell(self.index + 1, 6, result)  # 把对比的结果写入到表格中
        wb.save(self.filename)  # 保存


    def saveExcel(self):
        pass





# if __name__ == '__main__':
#     run = Write_excel().write("./Test.xlsx",1,"你好，摩西摩西")